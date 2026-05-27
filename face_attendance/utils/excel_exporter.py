"""
Excel导出模块
"""
import os
from datetime import datetime
from typing import List, Dict
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Excel导出器"""

    @staticmethod
    def export_attendance_records(records: List[Dict], file_path: str,
                                    title: str = "考勤记录") -> bool:
        """
        导出考勤记录到Excel

        Args:
            records: 考勤记录列表（字典列表，键为中文）
            file_path: 保存路径
            title: 表格标题

        Returns:
            是否成功
        """
        if not records:
            return False

        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "考勤记录"

            # 定义样式
            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            cell_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # 写入标题
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(records[0]))
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            # 写入表头
            headers = list(records[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # 写入数据
            for row_idx, record in enumerate(records, 4):
                for col_idx, header in enumerate(headers, 1):
                    value = record.get(header, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = cell_alignment
                    cell.border = thin_border

                    # 状态列着色
                    if header == "状态":
                        if value == "正常":
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        elif value == "迟到":
                            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        elif value == "缺勤":
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # 自动调整列宽
            for col_idx, header in enumerate(headers, 1):
                max_length = len(str(header))
                for record in records:
                    value = str(record.get(header, ""))
                    max_length = max(max_length, len(value))
                ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 4

            # 保存文件
            wb.save(file_path)
            return True

        except Exception as e:
            print(f"导出Excel失败: {e}")
            return False

    @staticmethod
    def export_student_list(students: List[Dict], file_path: str) -> bool:
        """
        导出学生名单到Excel

        Args:
            students: 学生列表
            file_path: 保存路径

        Returns:
            是否成功
        """
        if not students:
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "学生名单"

            # 样式
            header_font = Font(bold=True, size=11, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # 表头
            headers = ["学号", "姓名", "班级", "注册时间"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 数据
            for row_idx, student in enumerate(students, 2):
                ws.cell(row=row_idx, column=1, value=student.get('student_id', ''))
                ws.cell(row=row_idx, column=2, value=student.get('name', ''))
                ws.cell(row=row_idx, column=3, value=student.get('class_name', ''))
                created_at = student.get('created_at')
                if created_at:
                    ws.cell(row=row_idx, column=4, value=created_at.strftime('%Y-%m-%d %H:%M:%S'))

            # 调整列宽
            for col_idx in range(1, 5):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"导出学生名单失败: {e}")
            return False

    @staticmethod
    def generate_report(records: List[Dict], course_name: str,
                        start_date: str, end_date: str, file_path: str) -> bool:
        """
        生成考勤报告

        Args:
            records: 考勤记录
            course_name: 课程名称
            start_date: 开始日期
            end_date: 结束日期
            file_path: 保存路径

        Returns:
            是否成功
        """
        if not records:
            return False

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "考勤报告"

            # 标题
            ws.merge_cells('A1:H1')
            ws.cell(row=1, column=1, value=f"{course_name} 考勤报告")
            ws.cell(row=1, column=1).font = Font(bold=True, size=16)
            ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

            # 报告信息
            ws.cell(row=3, column=1, value=f"报告生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            ws.cell(row=4, column=1, value=f"统计时间段: {start_date} 至 {end_date}")

            # 统计数据
            total = len(records)
            normal = sum(1 for r in records if r.get('状态') == '正常')
            late = sum(1 for r in records if r.get('状态') == '迟到')

            ws.cell(row=6, column=1, value=f"总打卡次数: {total}")
            ws.cell(row=6, column=3, value=f"正常: {normal}")
            ws.cell(row=6, column=5, value=f"迟到: {late}")

            # 详细记录
            # ... 可扩展

            wb.save(file_path)
            return True

        except Exception as e:
            print(f"生成报告失败: {e}")
            return False